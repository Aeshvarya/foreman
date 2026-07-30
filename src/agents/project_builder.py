"""Build a whole project from a sentence, or from the spreadsheet you already have.

The guided form works, but it still asks a non-technical user to think like a
scheduler: enumerate suppliers, then materials, then activities, then wire the
dependencies by hand. Most people cannot start there. They *can* say:

    "40MW data centre in Chennai, handover 15 March. Steel from Tata, switchgear
     from Siemens, chillers from Blue Star. Steel lands mid-Jan."

…or drop the procurement tracker they already keep in Excel. Both are the same
problem — messy human input in, a valid project graph out — so both take the
same path: turn the input into text, ask the model for structured JSON, then
run it through the *existing* validation in `projects._normalise` so nothing
enters the engine that the guided builder could not have produced.

Nothing is saved here. This returns a DRAFT for the user to look over, because
a graph invented from a sentence must be confirmed by a human before it becomes
the thing their decisions rest on.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import projects                                  # noqa: E402
from agents.llm import has_key, invoke_text      # noqa: E402

MAX_INPUT_CHARS = 24_000

PROMPT = """You are a construction planner. Turn the description below into a \
structured project for a supply-chain scheduling tool.

TODAY IS {today}.

Return ONLY a JSON object, no prose, no markdown fences, with this exact shape:

{{
  "project": {{"name": "...", "start_date": "YYYY-MM-DD", "handover_milestone": "ACT-n"}},
  "suppliers":  [{{"id": "SUP-1", "name": "...", "region": "north|south|east|west|central", "reliability": 0.85}}],
  "materials":  [{{"id": "MAT-1", "name": "...", "supplier": "SUP-1",
                   "expected_arrival": "YYYY-MM-DD", "roj_date": "YYYY-MM-DD",
                   "confidence": 0.8, "lead_time_days": 30}}],
  "activities": [{{"id": "ACT-1", "name": "...", "duration_days": 10,
                   "needs_materials": ["MAT-1"], "depends_on": []}}]
}}

RULES
- Ids must be exactly SUP-n, MAT-n, ACT-n numbered from 1.
- Every material must name a supplier that exists in the suppliers list.
- `roj_date` is when the material is REQUIRED ON JOB; `expected_arrival` is when \
it is currently expected. If the user gives only one, set both the same.
- If dates are missing, infer sensible ones working BACK from the handover date \
and forward from the start date. Never leave a date blank.
- `depends_on` is the activity or activities that must FINISH first. Build a \
realistic construction sequence (foundations before structure before fit-out \
before testing before handover). No activity may depend on itself; no cycles.
- The LAST activity must be the handover/commissioning milestone, and \
`handover_milestone` must be its id.
- `confidence` is how sure we are of the material's date: 0.9+ if the user says \
confirmed/delivered, ~0.8 if stated plainly, ~0.6 if vague or unknown.
- If the description is thin, invent the minimum realistic structure to make a \
working schedule — but never invent a supplier or material the user did not \
imply.
- Aim for 3-12 activities. Keep names short and human ("Roof steel erection").

DESCRIPTION:
{text}
"""


def draft(text: str) -> dict:
    """Free text (or spreadsheet text) -> a validated draft project + summary."""
    text = (text or "").strip()
    if len(text) < 12:
        raise ValueError("Tell me a bit more about the project — even one sentence.")
    if not has_key():
        raise RuntimeError(
            "The describe-it option needs a Gemini key in .env. You can still "
            "build the project with the form."
        )
    text = text[:MAX_INPUT_CHARS]

    raw = invoke_text(PROMPT.format(today=date.today().isoformat(), text=text), 0.0)
    data = _parse_json(raw)
    if data is None:
        # One retry, blunter. Free-tier models occasionally wrap or chat.
        raw = invoke_text(
            "Return ONLY the JSON object, nothing else. No markdown.\n\n"
            + PROMPT.format(today=date.today().isoformat(), text=text), 0.0,
            use_cache=False)
        data = _parse_json(raw)
    if data is None:
        raise ValueError("Could not read a project out of that. Try describing it "
                         "in a couple of plain sentences.")

    warnings = _repair(data)

    # The draft goes through exactly the same validation as the guided builder —
    # dangling references dropped, cycles rejected, defaults backfilled.
    try:
        clean = projects._normalise(data)
    except (KeyError, ValueError) as e:
        raise ValueError(f"That came out as an invalid schedule ({e}). "
                         f"Try naming the main materials and the handover date.")

    return {
        "draft": clean,
        "summary": _summary(clean),
        "warnings": warnings,
        "counts": {
            "suppliers": len(clean["suppliers"]),
            "materials": len(clean["materials"]),
            "activities": len(clean["activities"]),
        },
    }


# ------------------------------------------------------------------ parsing
def _parse_json(raw: str) -> dict | None:
    """Models like to wrap JSON in prose or fences. Dig it out."""
    if not raw:
        return None
    raw = raw.strip()
    fence = re.search(r"```(?:json)?\s*(.+?)```", raw, re.S)
    if fence:
        raw = fence.group(1).strip()
    start, end = raw.find("{"), raw.rfind("}")
    if start == -1 or end <= start:
        return None
    try:
        obj = json.loads(raw[start:end + 1])
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, dict) and "project" in obj else None


def _repair(data: dict) -> list[str]:
    """Fix what a model reliably gets slightly wrong, and say what we changed.

    Silent repair is how a demo turns into a lie, so every fix becomes a line
    the user sees on the confirmation screen.
    """
    warns: list[str] = []
    today = date.today()

    proj = data.setdefault("project", {})
    if not str(proj.get("name", "")).strip():
        proj["name"] = "New project"
        warns.append("No project name was given, so we called it 'New project'.")
    if not _is_date(proj.get("start_date")):
        proj["start_date"] = today.isoformat()
        warns.append(f"No start date given — assumed today ({today.isoformat()}).")

    sups = {s.get("id") for s in data.get("suppliers", [])}
    for m in data.get("materials", []):
        if m.get("supplier") not in sups and sups:
            m["supplier"] = sorted(sups)[0]
            warns.append(f"{m.get('name', 'A material')} had no matching supplier — "
                         f"attached it to {m['supplier']}.")
        # Dates are the one thing we must never leave the model guessing about
        # silently, because the whole cascade is built on them.
        for field, fallback in (("expected_arrival", 30), ("roj_date", 30)):
            if not _is_date(m.get(field)):
                other = m.get("roj_date" if field == "expected_arrival" else "expected_arrival")
                m[field] = other if _is_date(other) else (today + timedelta(days=fallback)).isoformat()
                warns.append(f"{m.get('name', 'A material')} was missing a date — "
                             f"used {m[field]}. Change it if that's wrong.")

    for a in data.get("activities", []):
        try:
            a["duration_days"] = max(1, int(a.get("duration_days", 5)))
        except (TypeError, ValueError):
            a["duration_days"] = 5
        a["depends_on"] = [d for d in (a.get("depends_on") or []) if d != a.get("id")]
        a.setdefault("needs_materials", [])

    return warns


def _is_date(v) -> bool:
    try:
        date.fromisoformat(str(v))
        return True
    except (TypeError, ValueError):
        return False


def _summary(clean: dict) -> str:
    """One line a human can check at a glance before committing."""
    handover = clean["project"]["handover_milestone"]
    last = next((a["name"] for a in clean["activities"] if a["id"] == handover), handover)
    return (f"{len(clean['suppliers'])} suppliers, {len(clean['materials'])} materials and "
            f"{len(clean['activities'])} jobs, finishing with \"{last}\".")


# -------------------------------------------------------------- spreadsheets
def spreadsheet_to_text(filename: str, content: bytes) -> str:
    """Flatten a CSV/Excel export to plain text rows for the same extraction.

    Deliberately not a column-mapping parser: real procurement trackers have
    merged cells, notes columns and three header rows. Handing the model the
    rows as text copes with that far better than a rigid schema would.
    """
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt", ".tsv")):
        text = content.decode("utf-8", errors="replace")
        if name.endswith(".tsv"):
            return text
        rows = list(csv.reader(io.StringIO(text)))
        return "\n".join(" | ".join(c.strip() for c in r) for r in rows if any(c.strip() for c in r))

    if name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError("Excel support needs openpyxl — save the sheet as CSV instead.")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"--- sheet: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    out.append(" | ".join(cells))
        wb.close()
        return "\n".join(out)

    raise ValueError(f"{filename}: upload a CSV or Excel file, or just describe the project.")


if __name__ == "__main__":
    d = draft("40MW data centre in Chennai, handover 15 March 2027. Structural steel "
              "from Tata Projects arriving mid-January, 4000A switchgear from Siemens, "
              "chillers from Blue Star. Steel is confirmed, switchgear is not.")
    print(d["summary"])
    for w in d["warnings"]:
        print("  !", w)
    print(json.dumps(d["draft"], indent=1)[:900])
